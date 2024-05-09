from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponsePermanentRedirect
from .models import Magnicon_CCC_Process, Thomas_Process, Warshawsky_Process, \
                    Scaling_CCC_Process, \
                    document, search_standard_resistor
from django.core.files.storage import FileSystemStorage
from .forms import documentation_form, calibration_area_form, search_standard_resistor_form
from resistors.data_handler import delete_records
from django.views import View
import csv, uuid, random
from django.views.generic.edit import FormView
#from .admin import PostMagniconCCC
from django.shortcuts import get_object_or_404
from django.http import FileResponse
from django.db import connections
from django.db.utils import OperationalError
import xlsxwriter
import mysql.connector
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
# Save the workbook to a BytesIO buffer
from io import BytesIO
import openpyxl
from datetime import datetime

# Create your views here.
def process(request):
    return render(request, 'process.html')
    
def index(request):
    return render(request, 'index.html')

def home(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        # Authenticate
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "You are logged in...")
            return redirect('selectCalibrationArea')
        else:
            messages.success(request, "There was an error loggin in...")
            return redirect('home')
    return render(request, 'home.html', {})

def logout_user(request):
    logout(request)
    messages.success(request, "You are logged out...")
    return redirect('home')

def selectCalibrationArea(request):
    return render(request, 'selectCalibrationArea.html')

def upload(request):
    if request.method == 'POST':
        #delete_records(magnicon_ccc)
        #delete_records(thomas_dcc)
        print('My file:', request.FILES.getlist('mdss_data_file[]', False))
        mdss_data_form = calibration_area_form(request.POST, request.FILES)
        if mdss_data_form.is_valid() and request.FILES.getlist('mdss_data_file[]', False) != False:
            # if file is uploaded, then handle the file.
            f_meta = handle_uploaded_file(mdss_data_form, request.FILES.getlist('mdss_data_file[]'))
            context = {'fdata':f_meta}
            return render(request, 'upload.html', context)
    else:
        mdss_data_form = calibration_area_form()
    return render(request, 'upload.html')

def handle_uploaded_file(myform, myfile):
    s_list = []
    my_slist = []
    all_files = []
    all_files_size = []
    all_files_type = []
    db_conn = connections['default']
    cursor = db_conn.cursor()
    
    for files in myfile:
        all_files.append(files) # get the filename...
        all_files_size.append(files.size) # get the filesize...
        all_files_type.append(files.content_type) # get the file type
        query = """INSERT INTO resistors_filename (id, date_uploaded, uploaded_filename) VALUES (%s, %s, %s)"""
        cursor.execute(query, (str(random.randint(0, 65535)), datetime.now().strftime("%d%m%Y_%H%M%S"), files))
        # read file 
        for lines in files.__iter__():
            s_list.append(str(random.randint(0, 65535)))
            s_list.extend(((lines.decode('utf-8')).split('|')))
            my_slist.append(s_list)
            s_list = []
        #print (my_slist)
        for i in my_slist:
            if i[-2] == 'Magnicon CCC Process':
                print('In magnicon ccc process')
                myobj = Magnicon_CCC_Process(*i)
                myobj.save()
                #delete_records(magnicon_ccc)
            elif i[-2] == 'Thomas Process':
                print("In thomas process...")
                print(i)
                myobj = Thomas_Process(*i)
                myobj.save()
                #delete_records(thomas_dcc)
            elif i[-2] == 'Scaling CCC Process':
                myobj = Scaling_CCC_Process(*i)
                myobj.save()
            elif i[-2] == 'Warshawsky Process':
                myobj = Warshawsky_Process(*i)
                myobj.save()
            elif i[-2] == 'MI 6010C Process':
                myobj = MI_6010C_Process(*i)
                myobj.save()
            elif i[-2] == 'MI 6010B Process':
                myobj = MI_6010B_Process(*i)
                myobj.save()
        myform.save(commit = False)
    fmeta = []
    cursor.close()
    for fn, fs, ft in zip(all_files, all_files_size, all_files_type):
        fmeta.append([fn, fs, ft])
    return (fmeta)

def search(request):
    """
    Handle's client request for searching...
    """
    if request.method == 'POST':
        search_data_form = search_standard_resistor_form(request.POST)
        for field in search_data_form:
            print("Field Error:", field.name,  field.errors)
        if search_data_form.is_valid():
            search_data_form.save()
            serial = search_data_form.cleaned_data['serial']
            process_name = search_data_form.cleaned_data['process_name']
            service_id = search_data_form.cleaned_data['service_id']
            format = search_data_form.cleaned_data['format']
            #print(serial, process_name, service_id)
            if serial != '' or process_name != '' or service_id != '' or format != '':
                mydict = {'Serial': serial,
                          'Process': process_name,
                          'Service ID': service_id,
                          'Format': format,}
                response = fetch_data(mydict)
                return(response)
            else:
                return redirect('.')
                pass
            # Assuming you want to redirect or do something after successful form submission
            # Replace 'redirect_url' with the actual URL you want to redirect to.
            
    else:
        # Handle GET request, just render the empty form
        search_data_form = search_standard_resistor_form()
    return render(request, 'search.html', {'form': search_data_form})

def get_all_tables():
    query = "SHOW TABLES"
    db_conn = connections['default']
    cursor = db_conn.cursor()
    cursor.execute(query)
    tables = [table[0] for table in cursor.fetchall()]
    cursor.close()
    return(tables)

def fetch_data(mydict):
    db_conn = connections['default']
    keys = list(mydict.keys())
    search_params = list(mydict.values())
    keyword_1 = search_params[0]
    keyword_2 = search_params[1]
    keyword_3 = search_params[2]
    tables = get_all_tables()
    cursor = db_conn.cursor()
    results = []
    table_names = []
    for table in tables:
        if  table.startswith('resistors') and \
            table != ('resistors_search_standard_resistor') and \
            table != ('resistors_document'):
            i=1
            #print("Table name: ", table)
            cursor.execute(f"DESCRIBE {table}")
            columns = [column[0] for column in cursor.fetchall()]
            param = []
            for val in search_params:
                if val:
                    i+=1
                    param.append(f"%{val}%")
            #print("Column names: ", columns)
            param_final = param*len(columns)
            query = "SELECT * FROM {} WHERE ".format(table)
            conditions = []
            # Generate the mysql query...
            for column in columns:
                #for key in keys:
                    #if column.upper() == key:
                column  = '`' + column  + '`'
                if len(param) == 1:
                    conditions.append("{} LIKE %s".format(column))
                elif len(param) == 2:
                    conditions.append("{} LIKE %s".format(column) + " AND " + "{} LIKE %s".format(column))
            query += " OR ".join(conditions)
            param_final =  tuple(param_final)
            #param = (f"%{keyword_1}%", f"%{keyword_2}%", )*len(columns)
            #print(query)
            #print (param_final)
            # execute the mysql query and fetch the results...
            cursor.execute(query, param_final)
            table_results = cursor.fetchall()
            results.extend(table_results)
            #print (table_results)
            if table_results != ():
                table_names.append(table)
    header = [row[0] for row in cursor.description]
    cursor.close()
    #export_xlsxwriter(header, results, mydict, table_names)
    response = export_openpyxl(header, results, mydict, table_names)
    return response
    # Fetch results
    #results = cursor.fetchall()

def fetch_table_data(table_name):
    db_conn = connections['default']
    cursor = db_conn.cursor()
    cursor.execute('select * from ' + table_name)
    header = [row[0] for row in cursor.description]
    rows = cursor.fetchall()
    cursor.close()
    return header, rows

def export_xlsxwriter(header, rows, search_params, table_names):
    # Create an new Excel file and add a worksheet.
    if search_params['Format'] == '':
        search_params['Format'] = '.csv'
    if search_params['Format'] == '.csv' or search_params['Format'] == '.xlsx' or search_params['Format'] == '.xls':
        workbook = xlsxwriter.Workbook('C:/Users/ohm/Downloads/' + 'Standard Resistor' + search_params['Format'])
        worksheet = workbook.add_worksheet(search_params['Serial'])
        # Create style for cells
        #header_cell_format = workbook.add_format({'bold': True, 'border': True, 'bg_color': 'yellow'})
        #body_cell_format = workbook.add_format({'border': True})
        row_index = 0
        column_index = 0
        for tables in table_names:
            worksheet.write(row_index, column_index, tables)
            row_index += 1
            column_index += 1
            for column_name in header:
                worksheet.write(row_index, column_index, column_name)
                column_index += 1
            row_index += 1
            for row in rows:
                column_index = 0
                for column in row:
                    worksheet.write(row_index, column_index, column)
                    column_index += 1
                row_index += 1
            print(str(row_index) + ' rows written successfully to ' + workbook.filename)
        # Closing workbook
        workbook.close()

def export_openpyxl(header, rows, search_params, table_names):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
     # Create an new Excel file and add a worksheet.
    if search_params['Format'] == '':
        search_params['Format'] = '.csv'
    if search_params['Format'] == '.csv' or search_params['Format'] == '.xlsx' or search_params['Format'] == '.xls':
        row_index = 1
        column_index = 1
        for tables in table_names:
            worksheet.cell(row_index, column_index).value = tables
            row_index += 1
            column_index += 1
            for column_name in header:
                worksheet.cell(row_index, column_index).value = column_name
                column_index += 1
            row_index += 1
            for row in rows:
                column_index = 1
                for column in row:
                    worksheet.cell(row_index, column_index).value = column
                    column_index += 1
                row_index += 1
        #workbook.save()
        # Save the workbook to a BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        # Create the HttpResponse object with the appropriate MIME type
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Set content disposition to force browser to download file
        response['Content-Disposition'] = 'attachment; filename="StandardResistor.xlsx"'
    return response 

def calibrationArea(request):
    return render(request, 'calibrationArea.html')

def documentation(request):
    if request.method == 'POST':
        myform = documentation_form(request.POST, request.FILES)
        a = request.FILES
        if myform.is_valid():
            myform.save()
            return redirect('.')
    else:
        myform = documentation_form()
    return render(request, 'documentation.html', {'form': myform})