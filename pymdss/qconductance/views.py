from django.shortcuts import render
from time import sleep
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from .models import Quantum_Conductance_Process
from django.core.files.storage import FileSystemStorage
from .forms import calibration_area_form, search_qconductance_form
from resistors.data_handler import delete_records
from django.views import View
from django.views.generic.edit import FormView
from django.shortcuts import get_object_or_404
from django.db import connections
import mysql.connector
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
import django.apps
# Save the workbook to a BytesIO buffer
from io import BytesIO
import openpyxl
import datetime
from celery import shared_task, Celery
from celery_progress.backend import ProgressRecorder
from django.core.serializers import serialize, deserialize
from celery.result import AsyncResult
from zoneinfo import ZoneInfo
from time import sleep

import logging
logger = logging.getLogger(__name__)

celery_app = Celery('pymdss', backend="django-db")
global nrows

# Create your views here.
def index(request):
    return render(request, 'qconductance-index.html')

def qconductanceCalibrationArea(request):
    return render(request, 'qconductanceCalibrationArea.html')

def process(request):
        sleep(2)
        nrows = request.session.get('nrows', '0')
        return render(request, 'process.html', {'nrows': nrows})
    
def index(request):
    return render(request, 'index.html')

def home(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        # Authenticate
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "You are logged in...")
            return redirect('selectCalibrationArea')
        else:
            messages.success(request, "There was an error loggin in...")
            return redirect('home')
    return render(request, 'home.html', {})

def logout_user(request):
    logout(request)
    messages.success(request, "You are logged out...")
    return redirect('home')

def selectCalibrationArea(request):
    return render(request, 'selectCalibrationArea.html')

def upload(request):
    if request.method == 'POST':
        mdss_data_form = calibration_area_form(request.POST, request.FILES)
        if mdss_data_form.is_valid() and request.FILES.getlist('mdss_data_file[]', False) != False:
            # if file is uploaded, then handle the file.
            myfilelist = request.FILES.getlist('mdss_data_file[]')
            task = handle_uploaded_file.apply_async(args=[[files.read() for files in myfilelist], [files.name for files in myfilelist]], ignore_result=False)
            #print('Task status:', task.status)
            #print('Task ID:', task.id)
            mdss_data_form.save(commit = False)
            #context = {'fdata':f_meta}
            context = {'task_id': task.task_id,}
            return render(request, 'upload.html', context)
        else:
            return render(request, 'upload.html')
    else:
        mdss_data_form = calibration_area_form()
        return render(request, 'upload.html')

def get_task_status(request, task_id):
    task = AsyncResult(task_id)
    if task.state == 'PENDING':
        response = {
            'state': task.state,
            'status': 'Task is still processing...'
        }
    elif task.state == 'SUCCESS':
        response = {
            'state': task.state,
            'result': task.result  # This should be the list returned by the task
        }
    elif task.state == 'FAILURE':
        resffponse = {
            'state': task.state,
            'status': str(task.info)  # The error message if task failed
        }
    return JsonResponse(response)

@shared_task(bind=True, track_started=True)
def handle_uploaded_file(self, encoded_file_data, encoded_file_name):
    msg = []
    error_filenames = []
    uploaded_filenames = []
    already_processed_filenames = []
    db_conn = connections['default']
    cursor = db_conn.cursor()
    progress_recorder = ProgressRecorder(self)
    pass_flag = 1
    for ct, file_data in enumerate(encoded_file_data):
        # First check if the file was uploaded before...
        query  = """SELECT uploaded_filename FROM resistors_filename WHERE uploaded_filename = "{}" """.format(encoded_file_name[ct])
        cursor.execute(query)
        result = cursor.fetchall()
        num_rows = cursor.rowcount
        #print(result, num_rows)
        if num_rows <= 0:
            mylist = file_data.decode('utf-8').split('\n')
            #print('mylist', mylist)
            for data in mylist:
                try:
                    if data != '':
                        data = 'None|' + data
                        i = data.split('|')
                        if i[-1] == '' or i[-1] == '\r' or i[-1] == '\n' or i[-1] =='\r\n':
                            i.pop(-1)
                        if 'Magnicon CCC Process' in i:
                            if len(i) == 46:
                                i[2] = datetime.datetime.strptime(i[2], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                                try:   
                                    i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                                except Exception as e:
                                    i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %I:%M %p').replace(tzinfo=ZoneInfo("UTC"))
                                    pass 
                                myobj = Magnicon_CCC_Process(*i)
                                myobj.id = None
                                myobj.save()
                        elif 'Thomas Process' in i:
                                #print('In Thomas process')
                                try:
                                    i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                                except Exception as e:
                                    i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                    pass
                                myobj = Thomas_Process(*i)
                                myobj.id = None
                                myobj.save()
                                #delete_records(thomas_dcc)
                        elif 'Scaling CCC Process' in i:
                            i[3] = datetime.datetime.strptime(i[3], '%Y-%m-%d %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            myobj = Scaling_CCC_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'Warshawsky Process' in i:
                            try:
                                i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                            except Exception as e:
                                i[3] = datetime.datetime.strptime(i[3], '%m/%d/%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                pass
                            myobj = Warshawsky_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'MI 6010C Process' in i:
                            try:
                                i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            except Exception as e:
                                i[2] = datetime.datetime.strptime(i[2], '%Y-%m-%d %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                pass
                            i[3] = datetime.datetime.strptime(i[3], '%H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            print('Data', len(i), i)
                            myobj = MI_6010C_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'MI 6010B Process' in i:
                            i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            i[3] = datetime.datetime.strptime(i[3], '%H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            myobj = MI_6010B_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'MI 6010SW Process' in i:
                            if len(i) == 21:
                                try:
                                    i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                except Exception as e:
                                    i[2] = datetime.datetime.strptime(i[2], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                                    pass
                                i[3] = datetime.datetime.strptime(i[3], '%H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                myobj = MI_6010SW_Process(*i)
                                myobj.id = None
                                myobj.save()
                        elif 'MI 6000B Process' in i:
                            i[2] = datetime.datetime.strptime(i[2], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                            try:
                                i[3] = datetime.datetime.strptime(i[3], '%H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            except Exception as e:
                                i[3] = datetime.datetime.strptime(i[3], '%H:%M').replace(tzinfo=ZoneInfo("UTC"))
                                pass
                            myobj = MI_6000B_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'MI 6010Q Process' in i:
                            try:
                                i[2] = datetime.datetime.strptime(i[2], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                            except Exception as e:
                                i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                pass
                            myobj = MI_6010Q_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'MI 6020Q Process' in i:
                            try:
                                i[2] = datetime.datetime.strptime(i[2], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                            except Exception as e:
                                i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                                pass
                            i[3] = datetime.datetime.strptime(i[3], '%H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            myobj = MI_6020Q_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'NIST AAB Process' in i:
                            #i.pop(-1)
                            #print('In NIST AAB process')
                            i[2] = datetime.datetime.strptime(i[2], '%m-%d-%Y %H:%M:%S').replace(tzinfo=ZoneInfo("UTC"))
                            myobj = NIST_AAB_Process(*i)
                            myobj.id = None
                            myobj.save()
                        elif 'HR3100 Process' in i:
                            i[5] = datetime.datetime.strptime(i[5], '%m/%d/%Y %I:%M:%S %p').replace(tzinfo=ZoneInfo("UTC"))
                            myobj = HR3100_Process(*i)
                            myobj.id = None
                            myobj.save()
                except Exception as e:
                    print('Error: ', e)
                    pass_flag = 0
                    #msg = 'Error processing file: ' + encoded_file_name[ct]
                    #return(msg)
                    break
            if pass_flag:
                # add the filename to the resistors_filename database
                query = """INSERT INTO resistors_filename (date_uploaded, uploaded_filename) VALUES (%s, %s)"""
                cursor.execute(query, (datetime.datetime.now().strftime("%d%m%Y_%H%M%S"), encoded_file_name[ct]))
                uploaded_filenames.append(encoded_file_name[ct])
            else:
                error_filenames.append(encoded_file_name[ct])    
        else:
            already_processed_filenames.append(encoded_file_name[ct])       
        progress_recorder.set_progress(int(((ct+1)/len(encoded_file_data))*100), 100)
    for i in uploaded_filenames:
        _msg = "Successfully uploaded Files: " +  str(i)
        msg.append(_msg)
    for i in error_filenames:
        _msg = "Error Files that couldn't be uploaded: " +  str(i)
        msg.append(_msg)
    for i in already_processed_filenames:
        _msg = "Already processed Files: " + str(i)
        msg.append(_msg)
    cursor.close()
    #sort_by_date()
    return (msg)

def search(request):
    """
    Handle's client request for searching...
    """
    if request.method == 'POST':
        search_data_form = search_standard_resistor_form(request.POST)
        #for field in search_data_form:
            #print("Field Error:", field.name,  field.errors)
        if search_data_form.is_valid():
            search_data_form.save()
            serial = search_data_form.cleaned_data['serial']
            nominal = search_data_form.cleaned_data['nominal']
            process_name = search_data_form.cleaned_data['process_name']
            service_id = search_data_form.cleaned_data['service_id']
            format = search_data_form.cleaned_data['format']
            #print(serial, process_name, service_id)
            if serial != '' or nominal != None or process_name != '' or service_id != '':
                mydict = {'Serial': serial,
                          'Nominal': nominal,
                          'Process': process_name,
                          'Service ID': service_id,
                          'Format': format,
                         }
                response = fetch_data(request, mydict)
                return response
            else:
                return redirect('.')
                pass
            # Assuming you want to redirect or do something after successful form submission
            # Replace 'redirect_url' with the actual URL you want to redirect to.     
    else:
        # Handle GET request, just render the empty form
        search_data_form = search_standard_resistor_form()
    return render(request, 'search.html', {'form': search_data_form})

def sort_by_date():
    resistors_tables = get_all_tables()
    #print(resistors_tables)
    db_conn = connections['default']
    cursor = db_conn.cursor()
    for table in resistors_tables:
        #print(table)
        if table == 'resistors_magnicon_ccc_process':
            query = """SELECT * FROM `{}` ORDER BY STR_TO_DATE(Date, '%m/%d/%Y %h:%i:%S %p') ASC""".format(table)
            try:
                cursor.execute(query)
            except mysql.connector.Error as err:
                print(f"Error: {err}")
        if table == 'resistors_scaling_ccc_process':
            query = """SELECT * FROM `{}` ORDER BY STR_TO_DATE(Date, '%Y-%m-%d %H:%i:%S') ASC""".format(table)
            try:
                cursor.execute(query)
            except mysql.connector.Error as err:
                print(f"Error: {err}")
    cursor.close()

def get_all_tables():
    query = "SHOW TABLES"
    db_conn = connections['default']
    cursor = db_conn.cursor()
    cursor.execute(query)
    tables = [table[0] for table in cursor.fetchall()]
    cursor.close()
    return(tables)

def fetch_data(request, mydict):
    db_conn = connections['default']
    keys = list(mydict.keys())
    search_params = list(mydict.values())
    tables = get_all_tables()
    cursor = db_conn.cursor()
    header = []
    results = []
    table_names = []
    mycol = []
    nrows = 0
    for table in tables:
        if  table.startswith('resistors') and \
            table != ('resistors_search_standard_resistor') and \
            table != ('resistors_document') and table != ('resistors_filename'):
            i=1
            #print("Table name: ", table)
            cursor.execute(f"DESCRIBE {table}")
            columns = [column[0] for column in cursor.fetchall()] # column list of all models
            param = []
            for val in search_params:
                if val:
                    i+=1
                    param.append(f"{val}")
            #print("Column names: ", columns)
            param_final = param
            query = "SELECT * FROM {} WHERE ".format(table)
            conditions = []
            # Generate the mysql query...
            for column in columns:
                for key in keys:
                    if mydict[key]:
                        if column == key:
                            mycol.append(column)
            #print ('Column: ', mycol)
            if len(param) == 1:
                conditions.append("BINARY {}=%s".format(mycol[0]))
            elif len(param) == 2:
                conditions.append("{} LIKE %s".format(mycol[0]) + " AND " + "{} LIKE %s".format(mycol[1]))
            elif len(param) == 3:
                conditions.append("{} LIKE %s".format(mycol[0]) + " AND " + "{} LIKE %s".format(mycol[1]) + " AND " + "{} LIKE %s".format(mycol[2]))
            elif len(param) == 4:
                conditions.append("{} LIKE %s".format(mycol[0]) + " AND " + "{} LIKE %s".format(mycol[1]) + " AND " + "{} LIKE %s".format(mycol[2]) + " AND " + "{} LIKE %s".format(mycol[3]))
            query += " OR ".join(conditions)
            param_final =  tuple(param_final)
            #param = (f"%{keyword_1}%", f"%{keyword_2}%", )*len(columns)
            #print(query)
            #print (param_final)
            # execute the mysql query and fetch the results...
            cursor.execute(query, param_final)
            table_results = cursor.fetchall()
            nrows += len(table_results)
            #print (table_results)
            if table_results != ():
                results.append(table_results)
                table_names.append(table)
                header.append([row[0] for row in cursor.description])
    cursor.close()
    print("rows: ", nrows)
    request.session['nrows'] = str(nrows)
    #export_xlsxwriter(header, results, mydict, table_names)
    if search_params[-1] == 'xlsx':
        response = export_openpyxl(header, results, mydict, table_names)
    elif search_params[-1] == 'text':
        response = export_text(header, results, mydict, table_names)
    else:
        response = export_openpyxl(header, results, mydict, table_names)
    return response
    # Fetch results
    #results = cursor.fetchall()

def fetch_table_data(table_name):
    db_conn = connections['default']
    cursor = db_conn.cursor()
    cursor.execute('select * from ' + table_name)
    header = [row[0] for row in cursor.description]
    rows = cursor.fetchall()
    cursor.close()
    return header, rows

def export_openpyxl(header, rows, mydict, table_names):
    # Create a new Excel workbook
    """
    mydict = {'Serial': serial,
            'Nominal': nominal,
            'Process': process_name,
            'Service ID': service_id,
            }
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    search_params = list(mydict.values())
    serial = search_params[0]
    nominal = search_params[1]
    process = search_params[2]
    service_id = search_params[3]
    not_allowed = ['[', ']', '*', '/', '\\', '?', ':']
    if serial != '' and len(serial) <=31:
        for i in not_allowed:
            if i in serial:
                serial = serial.replace(i, '')
        worksheet.title = str(serial)
    else:
        worksheet.title = "pymdss"
    # Create an new Excel file and add a worksheet.
    if True: #search_params['Format'] == '.csv' or search_params['Format'] == '.xlsx' or search_params['Format'] == '.xls':
        row_index = 1
        column_index = 1
        for head, tables, row in zip(header, table_names, rows):
            #print(head, tables)
            table_string_list= tables.split('_')
            table_string_list.pop(0)
            table_ = ' '.join(table_string_list)
            worksheet.cell(row_index, column_index).value = table_
            row_index += 1
            for column_name in head:
                if column_name != 'id':
                    worksheet.cell(row_index, column_index).value = column_name
                    column_index += 1
            row_index += 1
            column_index = 1
            for r in row:
                for column in r[1:]:
                    worksheet.cell(row_index, column_index).value = column
                    column_index += 1
                row_index += 1
                column_index = 1
        #workbook.save()
        # Save the workbook to a BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        # Create the HttpResponse object with the appropriate MIME type
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Set content disposition to force browser to download file
        response['Content-Disposition'] = 'attachment; filename="StandardResistor.xlsx"'
    return response 

def export_text(header, rows, mydict, table_names):
    # Create a new Excel workbook
    """
    mydict = {'Serial': serial,
            'Nominal': nominal,
            'Process': process_name,
            'Service ID': service_id,
            }
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    search_params = list(mydict.values())
    serial = search_params[0]
    nominal = search_params[1]
    process = search_params[2]
    service_id = search_params[3]
    if serial != '':
         worksheet.title = str(serial)
    # Create an new Excel file and add a worksheet.
    if True: #search_params['Format'] == '.csv' or search_params['Format'] == '.xlsx' or search_params['Format'] == '.xls':
        row_index = 1
        column_index = 1
        for head, tables, row in zip(header, table_names, rows):
            #print(head, tables)
            worksheet.cell(row_index, column_index).value = tables
            row_index += 1
            for column_name in head:
                worksheet.cell(row_index, column_index).value = column_name
                column_index += 1
            row_index += 1
            column_index = 1
            for r in row:
                for column in r:
                    worksheet.cell(row_index, column_index).value = column
                    column_index += 1
                row_index += 1
                column_index = 1
        #workbook.save()
        # Save the workbook to a BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        # Create the HttpResponse object with the appropriate MIME type
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Set content disposition to force browser to download file
        response['Content-Disposition'] = 'attachment; filename="StandardResistor.txt"'
    return response 

def calibrationArea(request):
    return render(request, 'calibrationArea.html')

def documentation(request):
    num_rows = []
    model_name = []
    context = []
    if request.method == 'POST':
        myform = documentation_form(request.POST, request.FILES)
        a = request.FILES
        if myform.is_valid():
            myform.save()
            return redirect('.')
    else:
        all_models = django.apps.apps.get_models(include_auto_created=False, include_swapped=False)
        for m in all_models:
            num_rows.append(m.objects.count())
            model_name.append(m.__name__)
        for i, j in zip(model_name, num_rows):
            context.append({'model':i,
                            'rows':j})
        #c = Magnicon_CCC_Process.objects.count()
        #print (model_name, num_rows)
        myform = documentation_form()
    return render(request, 'documentation.html', {'form': myform, 
                                                  'context': context})